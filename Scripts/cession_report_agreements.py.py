import os

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
import django
django.setup()

from datetime import datetime
from retailer.models import Customer
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

start_date = datetime.strptime("2023-01-01", "%Y-%m-%d")
end_date = datetime.strptime("2023-12-31", "%Y-%m-%d")

from_date = "2023-01-01"
to_date = "2023-12-31"

Agreement_type = "ALL"

workbook = Workbook()
std = workbook['Sheet']
sheet = workbook.active
workbook.remove(std)
agreements_sheet = workbook.create_sheet('Agreements')

agreements_sheet.cell(row=1, column=5).value ="Cession Statement Reports"
cell = agreements_sheet.cell(row=1, column=5)
cell.font = Font(size=12, bold=True)

agreements_sheet.cell(row=2, column=1).value = 'Agreement Type'
agreements_sheet.cell(row=2, column=2).value = Agreement_type
cell = agreements_sheet.cell(row=2, column=1)
cell.font = Font(size=12, bold=True)

agreements_sheet.append([])
agreements_sheet.cell(row=6,column=4).value = "INCEPTION TO DATE"
cell = agreements_sheet.cell(row=6, column=4)
cell.font = Font(size=12, bold=True)

agreements_sheet.cell(row=7,column=1).value = "From"
agreements_sheet.cell(row=7,column=2).value = str(from_date)
cell = agreements_sheet.cell(row=7, column=1)
cell.font = Font(size=12, bold=True)

agreements_sheet.cell(row=8,column=1).value = "To"
agreements_sheet.cell(row=8,column=2).value = str(to_date)
cell = agreements_sheet.cell(row=8, column=1)
cell.font = Font(size=12, bold=True)

agreements_sheet.append([])

agreements_header = [
    "Agreement ID",
    "Date of Inception",
    "Total Premium",
    "Earned Premium",
    "Unearned Premium",
    "Percentage of Earned Premium",
]
agreements_sheet.append(agreements_header)
header_font = Font(size=12, bold=True)
for row in agreements_sheet.iter_rows(min_row=10, max_row=10):
    for cell in row:
        cell.font = header_font

agreements = Customer.objects.filter(purchased_date__range=(start_date, end_date))

for agreement in agreements:
    earned_premium = agreement.service.earned_reserved()
    unearned_premium = agreement.service.unearned_reserved()
    percentage_earned_premium = agreement.service.earned_reserved_percentage()
    total_premium = earned_premium + unearned_premium
    
    data = [
        agreement.agreement_number,
        agreement.purchased_date.strftime("%m/%d/%Y"),
        total_premium,
        earned_premium,
        unearned_premium,
        percentage_earned_premium,
    ]
    agreements_sheet.append(data)
    dealer_name = agreement.dealer.name
    underwriter_name = agreement.underwriter.name
    reinsurance_name = agreement.reinsurance.name

agreements_sheet.cell(row=3, column=1).value = 'Dealers'
agreements_sheet.cell(row=3, column=2).value = dealer_name
cell = agreements_sheet.cell(row=3, column=1)
cell.font = Font(size=12, bold=True)

agreements_sheet.cell(row=4, column=1).value ="Underwriters"
agreements_sheet.cell(row=4, column=2).value = underwriter_name
cell = agreements_sheet.cell(row=4, column=1)
cell.font = Font(size=12, bold=True)

agreements_sheet.cell(row=5, column=1).value ="Reinsurance"
agreements_sheet.cell(row=5, column=2).value = reinsurance_name
cell = agreements_sheet.cell(row=5, column=1)
cell.font = Font(size=12, bold=True)

workbook.save("./cession_reports_agreements.xlsx")
