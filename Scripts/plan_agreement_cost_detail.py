import os

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
import django
django.setup()

from openpyxl.styles import Font
from openpyxl import Workbook
from retailer.models import Customer

workbook = Workbook()
std = workbook['Sheet']
sheet = workbook.active
workbook.remove(std)

plan_sheet = workbook.create_sheet('Plan Agreements Cost Details')
plan_header = [
    "Customer Name",
    "Agreement Number",
    "Agreement Plan Cost",
    "Term Cost",
    "SubPlan Id",
    "SubPlan Name",
    "Term ID",
    "Term Name",
    "Term Obligor Fee",
    "Agreement Obligor Fee",
    "Term Premium tax",
    "Agreement Premium tax",
    "Term Overfunds",
    "Agreement Overfunds",
    "Term Misc Overfunds",
    "Agreement Misc Overfunds",
    "Term Clip fee",
    "Agreement Clip fee",
    "Term Roadside",
    "Agreement Roadside",
    "Term Admin fee",
    "Agreement Admin fee",
    "Term Reserve",
    "Agreement Reserve",
    "Term Agent commission",
    "Agreement Agent commission",


    ]
plan_sheet.append([])
row_index = 2
for i in range(1, len(plan_header) + 1):
    plan_sheet.cell(row=1, column=i).value = plan_header[i - 1]
    plan_sheet.cell(row=1, column=i).font = Font(size=12, bold=True)

plan_id = 52

agreements = Customer.objects.filter(rate_card__subplan__plan__id=plan_id)
for agreement in agreements:
    subplan = agreement.rate_card.subplan
    term = agreement.rate_card
    plan_data = [
        agreement.customer_name,
        agreement.agreement_number,
        agreement.plan_cost,
        term.cost,                     
        subplan.id,
        subplan.name,
        term.id,
        term.name,
        term.obligor_fee,
        agreement.get_data_field("obligor_fee"),
        term.premium_tax,
        agreement.get_data_field("premium_tax"),
        term.over_funds,
        agreement.get_data_field("over_funds"),
        term.misc_over_funds,
        agreement.get_data_field("misc_over_funds"),
        term.clip_fee,
        agreement.get_data_field("clip_fee"),
        term.roadside,
        agreement.get_data_field("roadside"),
        term.admin_fee,
        agreement.get_data_field("admin_fee"),
        term.reserve,
        agreement.get_data_field("reserve"),
        term.agent_commission,
        agreement.get_data_field("agent_commission"),           
    ]
    plan_sheet.append(plan_data)
    row_index += 1

workbook.save("./plan_agreement_cost_detail.xlsx")