import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
import django
django.setup()

from openpyxl.styles import Font
from openpyxl import Workbook
from company.models import Plan
from retailer.models import Customer

# Create a new workbook and sheets
workbook = Workbook()
std = workbook['Sheet']
workbook.remove(std)
plan_sheet = workbook.create_sheet('For Plan')
agreement_sheet = workbook.create_sheet('For Agreement')

# Set header for the plan sheet
plan_header = ["Plan ID","Plan Name","Dealer"]
plan_sheet.append(plan_header)
header_font = Font(size=12, bold=True)
for cell in plan_sheet[1:1]:
    cell.font = header_font

# For Plans
plans = Plan.objects.all()

# check for missing money movement
for plan in plans:
    if plan.enable_money_movement is True:
        if not plan.agent_commission_plan.all() or not plan.overfund_plan.all() or not plan.misc_overfund_plan.all() or not plan.admin_fee_plan.all():
            plan_sheet.append([plan.id, plan.name, plan.dealer.name])

# Set header for the agreement sheet
agreement_header = ["Agreement Number","Dealer"]
agreement_sheet.append(agreement_header)
header_font = Font(size=12, bold=True)
for cell in agreement_sheet[1:1]:
    cell.font = header_font

# For agreements 
agreements = Customer.objects.all()

# check for missing money movement
for agreement in agreements:
    if agreement.enable_money_movement is True:
        if not agreement.overfund_agreement.all() or not agreement.misc_overfund_agreement.all() or not agreement.agent_commission_agreement.all() or not agreement.admin_fee_agreement.all():
            agreement_sheet.append([agreement.agreement_number,agreement.dealer.name])

# Save the workbook
workbook.save("./missing_money_movement.xlsx")