import os
import django


os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
django.setup()

from openpyxl.styles import Font
from openpyxl import Workbook
from retailer.models import (
    Customer,
    OverfundDistributionAgreement,
    MiscOverfundDistributionAgreement,
    AgentCommissionDistributionAgreement,
    RoadsideDistributionAgreement,
    AdminFeeDistributionAgreement,
    ObligorFeeDistributionAgreement,
    ReinsuranceDistributionAgreement,
)

valid_member_choices = ["Admin", "Dealer", "SR", "Agency", "Seller", "Underwriter", "Reinsurance"]
agreement_number = "TX1002-10117"
agreements = Customer.objects.filter(agreement_number=agreement_number, enable_money_movement=True)

workbook = Workbook()
std = workbook['Sheet']
workbook.remove(std)
missing_distributions_sheet = workbook.create_sheet('Missing Distribution')

missing_distributions_header = ["Agreement ID", "Distributions Name", "Distributions ID ","Member of Name"]
missing_distributions_sheet.append(missing_distributions_header)

header_font = Font(size=12, bold=True)
for cell in missing_distributions_sheet[1]:
    cell.font = header_font

for agreement in agreements:
    overfund_agreements = OverfundDistributionAgreement.objects.filter(agreement=agreement, member__isnull=False)
    misoverfund_agreements = MiscOverfundDistributionAgreement.objects.filter(agreement=agreement, member__isnull=False)
    agent_comm_agreements = AgentCommissionDistributionAgreement.objects.filter(agreement=agreement, member__isnull=False)
    admin_fee_agreements = AdminFeeDistributionAgreement.objects.filter(agreement=agreement, member__isnull=False)
    road_side_agreements = RoadsideDistributionAgreement.objects.filter(agreement=agreement, vendor__isnull=False)
    obligor_fee_agreements = ObligorFeeDistributionAgreement.objects.filter(agreement=agreement, member__isnull=False)
    reinsurance_agreements = ReinsuranceDistributionAgreement.objects.filter(agreement=agreement, reinsurance__isnull=False)

    for overfund_agreement in overfund_agreements:
        if overfund_agreement.member_of == "Admin":
            missing_distributions_sheet.append([agreement.agreement_number, "OverfundDistributionAgreement", overfund_agreement.id, overfund_agreement.member_of])
        elif overfund_agreement.member_of == "Dealer":
            missing_distributions_sheet.append([agreement.agreement_number, "OverfundDistributionAgreement", overfund_agreement.id, overfund_agreement.member_of])
        elif overfund_agreement.member_of == "SR":
            missing_distributions_sheet.append([agreement.agreement_number, "OverfundDistributionAgreement", overfund_agreement.id, overfund_agreement.member_of])
        elif overfund_agreement.member_of == "Agency":
            missing_distributions_sheet.append([agreement.agreement_number, "OverfundDistributionAgreement", overfund_agreement.id, overfund_agreement.member_of])
        elif overfund_agreement.member_of == "Seller":
            missing_distributions_sheet.append([agreement.agreement_number, "OverfundDistributionAgreement", overfund_agreement.id, overfund_agreement.member_of])
            
    for misoverfund_agreement in misoverfund_agreements:
        if misoverfund_agreement.member_of == "Admin":
            missing_distributions_sheet.append([agreement.agreement_number, "MiscOverfundDistributionAgreement", misoverfund_agreement.id, misoverfund_agreement.member_of])
        elif misoverfund_agreement.member_of == "Dealer":
            missing_distributions_sheet.append([agreement.agreement_number, "MiscOverfundDistributionAgreement", misoverfund_agreement.id, misoverfund_agreement.member_of])
        elif misoverfund_agreement.member_of == "SR":
            missing_distributions_sheet.append([agreement.agreement_number, "MiscOverfundDistributionAgreement", misoverfund_agreement.id, misoverfund_agreement.member_of])
        elif misoverfund_agreement.member_of == "Agency":
            missing_distributions_sheet.append([agreement.agreement_number, "MiscOverfundDistributionAgreement", misoverfund_agreement.id, misoverfund_agreement.member_of])
        elif misoverfund_agreement.member_of == "Seller":
            missing_distributions_sheet.append([agreement.agreement_number, "MiscOverfundDistributionAgreement", misoverfund_agreement.id, misoverfund_agreement.member_of])
        elif misoverfund_agreement.member_of == "underwriter":
            missing_distributions_sheet.append([agreement.agreement_number, "MiscOverfundDistributionAgreement", misoverfund_agreement.id, misoverfund_agreement.underwriter.name])
        elif misoverfund_agreement.member_of == "reinsurance":
            missing_distributions_sheet.append([agreement.agreement_number, "MiscOverfundDistributionAgreement", misoverfund_agreement.id, misoverfund_agreement.reinsurance.name])

    for agent_comm_agreement in agent_comm_agreements:
        missing_distributions_sheet.append([agreement.id, "AgentCommissionDistributionAgreement", agent_comm_agreement.id, agent_comm_agreement.agency.name])

    for admin_fee_agreement in admin_fee_agreements:
        if admin_fee_agreement.member_of == "Admin":
            missing_distributions_sheet.append([agreement.id, "AdminFeeDistributionAgreement", admin_fee_agreement.id, admin_fee_agreement.member_of])
        elif admin_fee_agreement.member_of == "Dealer":
            missing_distributions_sheet.append([agreement.id, "AdminFeeDistributionAgreement", admin_fee_agreement.id, admin_fee_agreement.member_of])
        elif admin_fee_agreement.member_of == "SR":
            missing_distributions_sheet.append([agreement.id, "AdminFeeDistributionAgreement", admin_fee_agreement.id, admin_fee_agreement.member_of])
        elif admin_fee_agreement.member_of == "Agency":
            missing_distributions_sheet.append([agreement.id, "AdminFeeDistributionAgreement", admin_fee_agreement.id, admin_fee_agreement.member_of])
        elif admin_fee_agreement.member_of == "Seller":
            missing_distributions_sheet.append([agreement.id, "AdminFeeDistributionAgreement", admin_fee_agreement.id, admin_fee_agreement.member_of])

    for road_side_agreement in road_side_agreements:
        missing_distributions_sheet.append([agreement.id, "RoadsideDistributionAgreement", road_side_agreement.id, road_side_agreement.vendor.name])

    for obligor_fee_agreement in obligor_fee_agreements:
        if obligor_fee_agreement.member_of == "Admin":
            missing_distributions_sheet.append([agreement.id, "ObligorFeeDistributionAgreement", obligor_fee_agreement.id, obligor_fee_agreement.member_of])
        elif obligor_fee_agreement.member_of == "Dealer":
            missing_distributions_sheet.append([agreement.id, "ObligorFeeDistributionAgreement", obligor_fee_agreement.id, obligor_fee_agreement.member_of])
        elif obligor_fee_agreement.member_of == "SR":
            missing_distributions_sheet.append([agreement.id, "ObligorFeeDistributionAgreement", obligor_fee_agreement.id, obligor_fee_agreement.member_of])
        elif obligor_fee_agreement.member_of == "Agency":
            missing_distributions_sheet.append([agreement.id, "ObligorFeeDistributionAgreement", obligor_fee_agreement.id, obligor_fee_agreement.member_of])
        elif obligor_fee_agreement.member_of == "Seller":
            missing_distributions_sheet.append([agreement.id, "ObligorFeeDistributionAgreement", obligor_fee_agreement.id, obligor_fee_agreement.member_of])
        elif obligor_fee_agreement.member_of == "underwriter":
            missing_distributions_sheet.append([agreement.id, "ObligorFeeDistributionAgreement", obligor_fee_agreement.id, obligor_fee_agreement.member_of])
        elif obligor_fee_agreement.member_of == "reinsurance":
            missing_distributions_sheet.append([agreement.id, "ObligorFeeDistributionAgreement", obligor_fee_agreement.id, obligor_fee_agreement.member_of])

    for reinsurance_agreement in reinsurance_agreements:
        missing_distributions_sheet.append([agreement.id, "ReinsuranceDistributionAgreement", reinsurance_agreement.id,  reinsurance_agreement.reinsurance.name])


output_file = 'Missing_member_distributions.xlsx'
workbook.save(output_file)
print(f"successfully {output_file}")




# import os
# import django

# os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
# django.setup()

# from openpyxl.styles import Font
# from openpyxl import Workbook
# from retailer.models import (
#     Customer,
#     OverfundDistributionAgreement,
#     MiscOverfundDistributionAgreement,
#     AgentCommissionDistributionAgreement,
#     RoadsideDistributionAgreement,
#     AdminFeeDistributionAgreement,
#     ObligorFeeDistributionAgreement,
#     ReinsuranceDistributionAgreement,
# )

# # Define valid member choices
# valid_member_choices = ["Admin", "Dealer", "SR", "Agency", "Seller", "Underwriter", "Reinsurance"]

# # Filter agreements
# agreement_number = "TX1002-10117"
# agreements = Customer.objects.filter(agreement_number=agreement_number, enable_money_movement=True)

# # Create workbook and sheet
# workbook = Workbook()
# std = workbook['Sheet']
# workbook.remove(std)
# missing_distributions_sheet = workbook.create_sheet('Missing Distribution')

# # Add header to sheet
# missing_distributions_header = ["Agreement ID", "Distributions Name", "Distributions ID", "Member of Name"]
# missing_distributions_sheet.append(missing_distributions_header)

# # Style the header
# header_font = Font(size=12, bold=True)
# for cell in missing_distributions_sheet[1]:
#     cell.font = header_font

# # Function to convert complex objects to string
# def convert_to_str(value):
#     if value is None:
#         return ''
#     if isinstance(value, str):
#         return value
#     try:
#         return str(value)
#     except Exception as e:
#         return f"Error: {e}"

# # Iterate over agreements
# for agreement in agreements:
#     # Get existing members for various agreement types
#     existing_members = set()
    
#     existing_members.update(
#         OverfundDistributionAgreement.objects.filter(agreement=agreement).values_list('member_of', flat=True).distinct()
#     )
#     existing_members.update(
#         MiscOverfundDistributionAgreement.objects.filter(agreement=agreement).values_list('member_of', flat=True).distinct()
#     )
#     existing_members.update(
#         AgentCommissionDistributionAgreement.objects.filter(agreement=agreement).values_list('agency', flat=True).distinct()
#     )
#     existing_members.update(
#         AdminFeeDistributionAgreement.objects.filter(agreement=agreement).values_list('member_of', flat=True).distinct()
#     )
#     existing_members.update(
#         RoadsideDistributionAgreement.objects.filter(agreement=agreement).values_list('vendor', flat=True).distinct()
#     )
#     existing_members.update(
#         ObligorFeeDistributionAgreement.objects.filter(agreement=agreement).values_list('member_of', flat=True).distinct()
#     )
#     existing_members.update(
#         ReinsuranceDistributionAgreement.objects.filter(agreement=agreement).values_list('reinsurance', flat=True).distinct()
#     )
    
#     # Determine missing members
#     missing_members = [member for member in valid_member_choices if member not in existing_members]
    
#     for missing_member in missing_members:
#         missing_distributions_sheet.append([agreement.agreement_number, "Missing Distribution", "", missing_member])

#     # Optionally, include existing distributions
#     for model, field_name in [
#         (OverfundDistributionAgreement, 'member_of'),
#         (MiscOverfundDistributionAgreement, 'member_of'),
#         (AgentCommissionDistributionAgreement, 'agency'),
#         (AdminFeeDistributionAgreement, 'member_of'),
#         (RoadsideDistributionAgreement, 'vendor'),
#         (ObligorFeeDistributionAgreement, 'member_of'),
#         (ReinsuranceDistributionAgreement, 'reinsurance'),
#     ]:
#         for distribution in model.objects.filter(agreement=agreement):
#             # Use convert_to_str to ensure all values are strings
#             missing_distributions_sheet.append([
#                 agreement.agreement_number,
#                 model.__name__,  # Using model name as the distribution name
#                 distribution.id,
#                 convert_to_str(getattr(distribution, field_name, ''))  # Convert complex object to string
#             ])

# # Save the workbook
# output_file = 'Missing_member_distributions.xlsx'
# workbook.save(output_file)
# print(f"Successfully saved {output_file}")
