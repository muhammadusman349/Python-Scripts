import os

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
import django
django.setup()
from retailer.models import CustomerClaim
from account.models import Dealer
from payment.models import Payment
from company.utils import AGREEMENTSUBTYPECHOICES, AGREEMENTCHOICES
from payment import PayableType
from django.utils import timezone
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook

def claim_reports():
    dealer_id = 1
    dealer = Dealer.objects.get(id=dealer_id)
    claim_start_date = ""
    claim_end_date = ""
    approved_not_paids = CustomerClaim.objects.filter(dealer__id=dealer_id)

    workbook = Workbook()
    std = workbook['Sheet']
    workbook.remove(std)
    claim_sheet = workbook.create_sheet('Approved Not Paid')
    claim_sheet.cell(row=1, column=1).value = 'Pending Claims Report'
    claim_sheet.cell(row=1, column=1).font = Font(size=18, bold=True)
    claim_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
    # Merge cells for the header
    claim_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=22)

    claim_sheet.cell(row=2, column=1).value = 'APPROVED NOT PAID'
    claim_sheet.cell(row=2, column=1).font = Font(size=18, bold=True)
    claim_sheet.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
    # Merge cells for the header
    claim_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    claim_header = [
        "Dealership#",
        "ClaimDate",
        "Claim#",
        "SubPlan",
        "Vin",
        "Customer",
        "Parts Cost",
        "Claim Type",
        "Component",
        "Labour Hours",
        "Labour Cost",
        "Miles Driven",
        "Elapsed Time",
        "Repiar Done By",
        "Total Claim",
        "Net Cost",
        "Inspection fee",
        "Paid by",
        "Filed By",
    ]
    # claim_sheet.append([])
    for header_item in claim_header:
        claim_sheet.cell(row=8, column=claim_header.index(header_item) + 1, value=header_item).font = Font(size=12, bold=True)
    total_parts_cost = 0
    total_labour_hours = 0     
    total_labour_cost = 0     
    total_total_claim = 0
    total_net_cost = 0
        
    data = {}
    row_index = 22
    for claim in approved_not_paids:
        if claim.approve == True and :
            data_list = [
                    claim.dealer.name,
                    claim.date.strftime("%m/%d/%Y"),
                    claim.claim_number,
                    claim.agreement.rate_card.subplan.name,
                    claim.agreement.VIN,
                    claim.agreement.customer_name,
                    claim.cal_part_cost(),
                    claim.claim_type,
                    claim.get_type_of_claim_components,
                    claim.cal_labor_hour(),
                    claim.cal_labor_cost(),
                    claim.miles_driven(),
                    claim.elapsed_time(),
                    claim.repair_done_by(),
                    claim.total,
                    claim.net_total,
                    claim.inspection_fee,
                    claim.paid_by if claim.paid_by else "",
                    claim.filled_by.name if claim.filled_by else "",
            ]
                    
            total_parts_cost+=claim.cal_part_cost()
            total_labour_hours+=claim.cal_labor_hour()
            total_labour_cost+=claim.cal_labor_cost()
            total_total_claim+=claim.total
            total_net_cost+=claim.net_total

            claim_sheet.append(data_list)
            
            
            subplan = claim.agreement.SubPlan
            plan_subtype = claim.agreement.rate_card.subplan.plan_subtype

            if plan_subtype == AGREEMENTCHOICES.DEBT_CANCELLATION:
                plan_subtype = "DEBT"
            elif plan_subtype == AGREEMENTCHOICES.MAINTENANCE:
                plan_subtype = "MAINT"
            if subplan in data.keys():
                if plan_subtype not in data[subplan].keys():
                    data[subplan] = {plan_subtype: {}}
                    data[subplan][plan_subtype]["count"] = 1
                    data[subplan][plan_subtype]["amount"] = round(claim, 2)
                    data[subplan][plan_subtype]["part_cost"] = round(claim.cal_part_cost(), 2)
                    data[subplan][plan_subtype]["labour_time"] = round(claim.cal_labor_hour(), 2)
                    data[subplan][plan_subtype]["miles_lapsed"] = int(claim.miles_driven())
                    data[subplan][plan_subtype]["time_lapsed"] = int(claim.elapsed_time())
                else:
                    data[subplan][plan_subtype]["count"] = data[subplan][plan_subtype]["count"] + 1
                    data[subplan][plan_subtype]["amount"] = round(data[subplan][plan_subtype]["amount"] + claim.total, 2)
                    data[subplan][plan_subtype]["part_cost"] = round(data[subplan][plan_subtype]["part_cost"] + claim.cal_part_cost(), 2)
                    data[subplan][plan_subtype]["labour_time"] = round(data[subplan][plan_subtype]["labour_time"] + claim.cal_labor_hour(), 2)
                    data[subplan][plan_subtype]["miles_lapsed"] = data[subplan][plan_subtype]["miles_lapsed"] + int(claim.miles_driven())
                    data[subplan][plan_subtype]["time_lapsed"] = data[subplan][plan_subtype]["time_lapsed"] + int(claim.elapsed_time())
                    
            else:
                data[subplan] = {plan_subtype: {}}
                data[subplan][plan_subtype] = {
                    "count": 1,
                    "amount": round(claim.total, 2),
                    "part_cost": round(claim.cal_part_cost(), 2),
                    "labour_time": round(claim.cal_labor_hour(), 2),
                    "miles_lapsed": int(claim.miles_driven()),
                    "time_lapsed": int(claim.elapsed_time()),
                    }
        print("data",data)
        claim_sheet.cell(row=3,column=2).value = 'Company'
        claim_sheet.cell(row=3,column=3).value = claim.dealer.company.name
        claim_sheet.cell(row=4,column=2).value = 'Dealer'
        claim_sheet.cell(row=4,column=3).value = claim.dealer.name
        claim_sheet.cell(row=4,column=4).value = claim.dealer.dealer_number
        claim_sheet.cell(row=5,column=2).value = 'Claim Start'
        claim_sheet.cell(row=5,column=3).value = "07/01/2023" #claim.agreement.purchased_date.strftime("%m/%d/%Y")
        claim_sheet.cell(row=6,column=2).value = 'Claim End'
        claim_sheet.cell(row=6,column=3).value = "07/01/2023" #claim.agreement.expiry_date.strftime("%m/%d/%Y")
        claim_sheet.cell(row=7,column=2).value = 'Created at'
        claim_sheet.cell(row=7,column=3).value = timezone.now().strftime("%m/%d/%Y")

            
    claim_sheet.append(["Total","","","","","", total_parts_cost,"","", total_labour_hours, total_labour_cost,"","","", total_total_claim, total_net_cost])
    for row in claim_sheet.iter_rows(min_row=claim_sheet.max_row, max_row=claim_sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top',wrap_text=True)

        
    claim_sheet.cell(row=row_index+1, column=1).value = 'Summary'
    claim_sheet.cell(row=row_index+1, column=1).font = Font(size=18, bold=True)
    claim_sheet.cell(row=row_index+1, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    claim_sheet.merge_cells(start_row=row_index+1, start_column=1, end_row=row_index+1, end_column=20)
    column = 1
    header_dict = []
    row_header_index = row_index
    for key, value in data.items():
        summary_headers = []
        print("key = {}...value = {}..".format(key, value))
        for i in value:
            claim_sheet.cell(row=row_index+2, column=column).value = i
            claim_sheet.cell(row=row_index+2, column=column+1).value = "Count"
            claim_sheet.cell(row=row_index+2, column=column+2).value = "Amount"
        # Ensure count_amount_data is a dictionary
        count_amount_data = value[next(iter(value))] if isinstance(value[next(iter(value))], dict) else {}
        print("Count and Amount data:", count_amount_data)
        claim_sheet.cell(row=row_header_index+3, column=1).value = key
        claim_sheet.cell(row=row_header_index+3, column=2).value = count_amount_data.get('count', 'No Count')
        claim_sheet.cell(row=row_header_index+3, column=3).value = count_amount_data.get('amount', 'No Amount')
        row_header_index += 1
        column +=3
    claim_sheet.cell(row=row_index+5, column=1).value = 'Averages'
    claim_sheet.cell(row=row_index+5, column=1).font = Font(size=10, bold=True)
    claim_sheet.cell(row=row_index+5, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    claim_sheet.merge_cells(start_row=row_index+5, start_column=1, end_row=row_index+5, end_column=5)
    count_amount_data = value[next(iter(value))] if isinstance(value[next(iter(value))], dict) else {}
    print("Count and Amount data:", count_amount_data)
    # claim_sheet.cell(row=row_header_index+3, column=1).value = key
    claim_sheet.cell(row=row_header_index+3, column=1).value = count_amount_data.get('part_cost', 'No Amount')
    claim_sheet.cell(row=row_header_index+4, column=1).value = count_amount_data.get('labour_time', 'No Amount')
    claim_sheet.cell(row=row_header_index+5, column=1).value = count_amount_data.get('miles_lapsed', 'No Amount')
    claim_sheet.cell(row=row_header_index+6, column=1).value = count_amount_data.get('time_lapsed', 'No Amount')

    claim_sheet.cell(row=row_index+10, column=1).value = 'Early Claims'
    claim_sheet.cell(row=row_index+10, column=1).font = Font(size=10, bold=True)
    claim_sheet.cell(row=row_index+10, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    claim_sheet.merge_cells(start_row=row_index+10, start_column=1, end_row=row_index+10, end_column=5)

    workbook.save("./claim_report.xlsx")
claim_reports()