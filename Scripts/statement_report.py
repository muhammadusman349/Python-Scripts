import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
import django
django.setup()
from company.models import Plan, Report, Company
from retailer.models import (
    Customer,
    MiscOverfundDistributionAgreement,
    )
from django.db import transaction
from payment.models import Payment,PayableType
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter,landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Frame,Table, TableStyle, NextPageTemplate, PageTemplate, BaseDocTemplate,PageBreak
from reportlab.lib.units import inch
from reportlab.lib.styles import ParagraphStyle

from django.core.files.uploadedfile import InMemoryUploadedFile
from datetime import datetime,date,timedelta
from account.models import Dealer,User
from company.models import Report
from conf.celery import app
from payment import PayableType
from  retailer.models import *
from io import BytesIO
from api.utils import check_last_quarter, quarter_cal
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
from payment.models import *

   

dealer_id = 3
workbook = Workbook()
std = workbook['Sheet']
workbook.remove(std)
remit_sheet = workbook.create_sheet('Remit Report')

created_at = datetime.now().date().strftime("%m/%d/%Y")

remit_sheet.cell(row=1, column=9).value = "Remit Agreement Payment Report"
bold_font = Font(bold=True, size=15)
# Apply the bold font to the cell
remit_sheet.cell(row=1, column=9).font = bold_font

remit_sheet.cell(row=2, column=2).value = 'Company'  
remit_sheet.cell(row=3, column=2).value = "Dealer"
remit_sheet.cell(row=4, column=2).value = "Period"
remit_sheet.cell(row=5, column=2).value = "Remit start"
remit_sheet.cell(row=6, column=2).value = "Remit end"
remit_sheet.cell(row=7, column=2).value = "Created At"
remit_sheet.cell(row=7, column=3).value = created_at
remit_sheet.append([])

remit_header = [
    "#",
    'Agreement#',
    'Remit Date',
    'Vin',
    'C Name',
    'Plan Type',
    'Admin Fee',
    'OverFund',
    'Agent Comm',
    'Reserve',
    'Premium Tax',
    'CLIP',
    'Intended Month',
    'Invoice',
    'Paid',
    'Balance',
]

remit_sheet.append(remit_header)
header_font = Font(size=12, bold=True)
for cell in remit_sheet[9]:
    cell.font = header_font

agreement_qs = AgreementPayment.objects.filter(agreement__dealer__id=dealer_id).order_by('id')
qs = agreement_qs.only('date', 'payment', 'paid_amount', 'agreement__agreement_number', 'agreement__term',
                       'agreement__VIN', 'agreement__recurring', 'agreement__middle_name',
                       'agreement__first_name', 'agreement__last_name', 'agreement__data',
                       'agreement__total_surcharge', 'agreement__plan_cost')
reserve_total = 0
tax_total = 0
clip_total = 0
dealer_payable_total = 0
paid_total = 0
balance_total = 0
agent_comm_total = 0
admin_total = 0
over_fund_total = 0
customer_ids=[]

for i in qs:
    customer_ids.append(i)
    if i.agreement.recurring:
        remit_data = [
            i.payment.id if i.payment else '',
            i.agreement.agreement_number,
            i.date.date().strftime("%m/%d/%Y") if i.date else "",
            i.agreement.VIN[-7:],
            (i.agreement.customer_name),
            (i.agreement.term),
            i.get_detail_format("admin_fee"),
            i.get_detail_format("over_funds"),
            i.get_detail_format("agent_commission"),
            round(i.get_detail_format("reserve") + i.get_detail_format("total_surcharge"), 2),
            i.get_detail_format("premium_tax"),
            i.get_detail_format("clip_fee"),
            (f"{i.get_detail_format('months')}"),
            i.agreement.dealer_payable(),
            i.agreement.paid_amount_value,
            i.agreement.balance()]
        remit_sheet.append(remit_data)
        admin_total += i.get_detail_format("admin_fee")
        over_fund_total += i.get_detail_format("over_funds")
        reserve_total += round(i.get_detail_format("reserve") + i.get_detail_format("total_surcharge"), 2)
        tax_total += round(i.get_detail_format("premium_tax"), 2)
        clip_total += round(i.get_detail_format("clip_fee"), 2)
        agent_comm_total += i.get_detail_format("agent_commission")

    else:
        rem_data = [
            i.payment.id if i.payment else '',
            i.agreement.agreement_number,
            i.date.date().strftime("%m/%d/%Y") if i.date else "",
            i.agreement.VIN[-7:],
            i.agreement.customer_name,
            i.agreement.term,
            i.get_data_field(field_name="admin_fee"),
            i.get_data_field(field_name="over_funds"),
            i.get_data_field(field_name="agent_commission"),
            round(i.get_data_field(field_name="reserve") + float(i.agreement.total_surcharge), 2),
            i.get_data_field(field_name="premium_tax"),
            i.get_data_field(field_name="clip_fee"),
            ("Full"),
            i.agreement.dealer_payable(),
            i.paid_amount,
            i.agreement.balance()
        ]
        remit_sheet.append(rem_data)
        
        admin_total += i.get_data_field(field_name="admin_fee")
        over_fund_total += float("{:.2f}".format(i.get_data_field(field_name="over_funds")))
        reserve_total += round(float(i.get_data_field(field_name="reserve")) + float(i.agreement.total_surcharge),2)
        tax_total += float("{:.2f}".format(i.get_data_field(field_name="premium_tax"))) if i.get_data_field(field_name="premium_tax") else 0
        clip_total += float("{:.2f}".format(i.get_data_field(field_name="clip_fee")))
        agent_comm_total += float("{:.2f}".format(i.get_data_field(field_name="agent_commission")))
    dealer_payable_total += float("{:.2f}".format(i.agreement.dealer_payable()))
    paid_total += float("{:.2f}".format(i.paid_amount))
    balance_total += float("{:.2f}".format(i.agreement.balance()))

remit_sheet.append([])
remit_sheet.append(["", "", "", "","","","","","Total"])

appended_row_number = remit_sheet.max_row
num_columns = len(remit_sheet[appended_row_number])
bold_font = Font(bold=True, size=15)
# Apply the bold font and font size to each cell in the appended row
for col_num in range(1, num_columns + 1):
    cell = remit_sheet.cell(row=appended_row_number, column=col_num)
    cell.font = bold_font

total_header = ['Type','Amount']

remit_sheet.append(total_header)
appended_row_number = remit_sheet.max_row
bold_font = Font(bold=True)

# Apply the bold font to each cell in the header row
for col_num in range(1, len(total_header) + 1):
    cell = remit_sheet.cell(row=appended_row_number, column=col_num)
    cell.font = bold_font   
     
totals = [
    ['Admin ', round(admin_total, 2)],
    ['Over Fund', round(over_fund_total, 2)],
    ['Agent Commission', round(agent_comm_total, 2)],
    ['Reserve', round(reserve_total, 2)],
    ['Premium Tax', round(tax_total, 2)],
    ['CLIP', round(clip_total, 2)],
    ['Dealer Invoice', round(dealer_payable_total, 2)],
    ['Paid', round(paid_total, 2)],
    ['Balance', round(balance_total, 2)]
]


# Iterate through the list of lists and append each to remit_sheet
for item in totals:
    remit_sheet.append(item)
    
remit_sheet.append([])
remit_sheet.append(["","","","","","","","","Summary"])

appended_row_number = remit_sheet.max_row
num_columns = len(remit_sheet[appended_row_number])
bold_font = Font(bold=True, size=15)
# Apply the bold font and font size to each cell in the appended row
for col_num in range(1, num_columns + 1):
    cell = remit_sheet.cell(row=appended_row_number, column=col_num)
    cell.font = bold_font
    
def just_optimized_local(ids_li=[]):
    data = {}
    for i in ids_li:
        term_name = i.agreement.term.lower()
        if term_name in data.keys():
            data[term_name]["count"] += 1
            data[term_name]["amount"] += round(i.paid_amount, 2)
        else:
            data[term_name] = {"count": 1, "amount": round(i.paid_amount, 2)}
    return data

summary_data = just_optimized_local(ids_li=customer_ids)

# Creating the summary header
summary_header = ['Type', 'Count', 'Amount']

remit_sheet.append(summary_header)
appended_row_number = remit_sheet.max_row
bold_font = Font(bold=True)

# Apply the bold font to each cell in the header row
for col_num in range(1, len(summary_header) + 1):
    cell = remit_sheet.cell(row=appended_row_number, column=col_num)
    cell.font = bold_font
# Adding data rows
for term_name, values in summary_data.items():
    remit_sheet.append([term_name.capitalize(), values['count'], values['amount']])
    
workbook.save("./remit_report.xlsx")
