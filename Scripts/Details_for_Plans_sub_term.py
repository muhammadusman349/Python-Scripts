import os

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
import django
django.setup()

from openpyxl.styles import Font
from openpyxl import Workbook
from company.models import Plan, SubPlan, Terms
from retailer.models import Customer, CustomerClaim

# Create a new workbook and sheets
workbook = Workbook()
std = workbook['Sheet']
workbook.remove(std)
plan_sheet = workbook.create_sheet('Plan')
subplan_sheet = workbook.create_sheet('SubPlan')
term_sheet = workbook.create_sheet('Term')
agreement_sheet = workbook.create_sheet('Agreement')
claim_sheet = workbook.create_sheet('Claim')


# Set header for the plan sheet
plan_header = ["Plan ID","Plan Name","Dealer"]
plan_sheet.append(plan_header)
header_font = Font(size=12, bold=True)
for cell in plan_sheet[1:1]:
    cell.font = header_font

# For Plans
plans = Plan.objects.all()
for plan in plans:
    plan_sheet.append([plan.id, plan.name, plan.dealer.name])

# Set header for the subplan sheet
subplan_header = ["SubPlan ID","SubPlan Name","Agreement Template"]
subplan_sheet.append(subplan_header)
header_font = Font(size=12, bold=True)
for cell in subplan_sheet[1:1]:
    cell.font = header_font

# For subPlans
subplans = SubPlan.objects.all()
for suplan in subplans:
    subplan_sheet.append([suplan.id, suplan.name, suplan.agreement_template.name])

# Set header for the term sheet
term_header = ["Term ID","Term Name","Type Of Claim"]
term_sheet.append(term_header)
header_font = Font(size=12, bold=True)
for cell in term_sheet[1:1]:
    cell.font = header_font

# For Terms
terms = Terms.objects.all()
for term in terms:
    term_sheet.append([term.id, term.name, term.subplan.name])

# Set header for the agreement sheet
agreement_header = ["Agreement Number","Dealer"]
agreement_sheet.append(agreement_header)
header_font = Font(size=12, bold=True)
for cell in agreement_sheet[1:1]:
    cell.font = header_font

# For agreements 
agreements = Customer.objects.all()
for agreement in agreements:
            agreement_sheet.append([agreement.agreement_number,agreement.dealer.name])

# Set header for the claim sheet
claim_header = ["Claim Number","Agreement","Dealer"]
claim_sheet.append(claim_header)
header_font = Font(size=12, bold=True)
for cell in claim_sheet[1:1]:
    cell.font = header_font

# For Claims 
claims = CustomerClaim.objects.all()
for claim in claims:
            claim_sheet.append([claim.claim_number,claim.agreement.agreement_number,claim.dealer.name])

# Save the workbook
workbook.save("./Plan_sub_term_details.xlsx")