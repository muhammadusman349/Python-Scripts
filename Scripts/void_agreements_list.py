import os

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
import django
django.setup()

from openpyxl.styles import Font
from openpyxl import Workbook

workbook = Workbook()
std = workbook['Sheet']
sheet = workbook.active
workbook.remove(std)
void_agreements_sheet = workbook.create_sheet('Void Agreements')
cancel_agreements_sheet = workbook.create_sheet('Cancel Agreements')

voided_agreements =  ['TX1021-28970', 'TX1021-33445', 'TX1021-19338', 'TX1021-17941', 'TX1021-29239', 'TX1021-34792', 'TX1021-31430', 'TX1021-29785',
                       'TX1021-25057', 'TX1021-31948-1', 'TX1021-25604', 'TX1021-27429', 'TX1021-20579', 'TX1021-32067-1', 'TX1021-32927', 'TX1021-27277', 'TX1021-29518', 'TX1021-22712', 'TX1021-18522', 
                       'TX1021-23142', 'TX1021-19133', 'TX1021-16934', 'TX1021-21301', 'TX1021-23712', 'TX1021-27169', 'TX1021-28802-1', 'TX1021-27770', 
                       'TX1021-36634', 'TX1021-30730', 'TX1021-30649', 'TX1021-28938', 'TX1021-32065', 'TX1021-21478', 'TX1021-36409', 'TX1021-29240', 
                       'TX1021-28032', 'TX1021-19043', 'TX1021-17039', 'TX1021-23778', 'TX1021-29009', 'TX1021-29224', 'TX1021-28873', 'TX1021-20688', 
                       'TX1021-17964', 'TX1021-29802', 'TX1021-30856', 'TX1021-24942', 'TX1021-21545', 'TX1021-22688', 'TX1021-31223-1', 'TX1021-33999', 
                       'TX1021-24439', 'TX1021-27569', 'TX1021-30274', 'TX1021-19050', 'TX1021-28176', 'TX1021-34036', 'TX1021-31294', 'TX1021-35473', 
                       'TX1021-22789', 'TX1021-27535', 'TX1021-29720', 'TX1021-30304-1', 'TX1021-30731', 'TX1021-28293', 'TX1021-19064', 'TX1021-29936', 
                       'TX1021-18341', 'TX1021-23217', 'TX1021-24957', 'TX1021-30947', 'TX1021-28861-1', 'TX1021-29926', 'TX1021-23311', 'TX1021-33685', 
                       'TX1021-19590', 'TX1021-19098', 'TX1021-24659', 'TX1021-29984', 'TX1021-17592', 'TX1021-18964', 'TX1021-32974', 'TX1021-16245', 
                       'TX1021-19402', 'TX1021-18189', 'TX1021-31222', 'TX1021-24958', 'TX1021-27446', 'TX1021-30994', 'TX1021-28711', 'TX1021-20197', 
                       'TX1021-30855', 'TX1021-24774', 'TX1021-23504', 'TX1021-29790', 'TX1021-35835-1', 'TX1021-27981', 'TX1021-30836', 'TX1021-33794-1', 
                       'TX1021-24340', 'TX1021-29784', 'TX1021-24016', 'TX1021-36893', 'TX1021-24773', 'TX1021-28942', 'TX1021-32420', 'TX1021-23682', 
                       'TX1021-29006', 'TX1021-29610', 'TX1021-27613', 'TX1021-30946', 'TX1021-35834', 'TX1021-17135', 'TX1021-29504', 'TX1021-23247', 
                       'TX1021-28292', 'TX1021-29221', 'TX1021-28859', 'TX1021-23974', 'TX1021-36506', 'TX1021-31862', 'TX1021-35108', 'TX1021-35926', 
                       'TX1021-29086', 'TX1021-32419', 'TX1021-32755', 'TX1021-28860', 'TX1021-20780', 'TX1021-18523', 'TX1021-36547', 'TX1021-35393', 
                       'TX1021-30993', 'TX1021-30000', 'TX1021-30956', 'TX1021-19065', 'TX1021-23777', 'TX1021-18188', 'TX1021-28943', 'TX1021-18278', 
                       'TX1021-36730', 'TX1021-32547', 'TX1021-30878', 'TX1021-21303', 'TX1021-33056', 'TX1021-27892',
                       'TX1021-27447', 'TX1021-28049', 'TX1021-28380', 'TX1021-29019', 'TX1021-30835', 'TX1021-18442', 'TX1021-34939', 
                       'TX1021-21092', 'TX1021-30857-1', 'TX1021-16672', 'TX1021-23363', 'TX1021-28132', 'TX1021-30350', 'TX1021-30578', 
                       'TX1021-32907', 'TX1021-24076', 'TX1021-29735', 'TX1021-24886-1', 'TX1021-35843', 'TX1021-31942', 'TX1021-36400', 
                       'TX1021-29277', 'TX1021-27814', 'TX1021-23367', 'TX1021-28971-1', 'TX1021-20689', 'TX1021-22846', 'TX1021-29644',
                         'TX1021-30577', 'TX1021-31290', 'TX1021-24809', 'TX1021-30524', 'TX1021-24658', 'TX1021-29989', 'TX1021-35504', 
                         'TX1021-19042', 'TX1021-32574', 'TX1021-29408', 'TX1021-27554', 'TX1021-34589', 'TX1021-28713', 'TX1021-31493', 
                         'TX1021-24438', 'TX1021-24112', 'TX1021-24075', 'TX1021-24004', 'TX1021-27276', 'TX1021-23713', 'TX1021-17564', 
                         'TX1021-29085', 'TX1021-27609', 'TX1021-29178', 'TX1021-29763', 'TX1021-28800', 'TX1021-34963-1', 'TX1021-31431', 
                         'TX1021-27536', 'TX1021-23366', 'TX1021-31748', 'TX1021-30081', 
                       'TX1021-31020', 'TX1021-30273', 'TX1021-25776', 'TX1021-32447', 'TX1021-23047', 'TX1021-29091', 
                       'TX1021-27430', 'TX1021-36464', 'TX1021-24148', 'TX1021-25191', 'TX1021-27652', 'TX1021-16357',
                        'TX1021-33446', 'TX1021-23246', 'TX1021-23719', 'TX1021-36086', 'TX1021-24113', 'TX1021-24017', 
                        'TX1021-27980', 'TX1021-21093', 'TX1021-36340', 'TX1021-21636', 'TX1021-18601', 'TX1021-30692', 
                        'TX1021-31943', 'TX1021-35990', 'TX1021-37021', 'TX1021-28937', 'TX1021-29460', 'TX1021-27188', 
                        'TX1021-29609', 'TX1021-24640', 'TX1021-17623', 'TX1021-23362', 'TX1021-23386', 'TX1021-27094', 'TX1021-20848', 
                        'TX1021-20356', 'TX1021-28632', 'TX1021-25058', 'TX1021-25212', 'TX1021-30879', 'TX1021-27392', 'TX1021-35023-1', 
                        'TX1021-29803', 'TX1021-20759', 'TX1021-19403', 'TX1021-30518', 'TX1021-36844', 'TX1021-29005', 'TX1021-27170', 'TX1021-23440', 
                        'TX1021-33018', 'TX1021-23439', 'TX1021-27555', 'TX1021-28747', 'TX1021-30202', 'TX1021-33816', 'TX1021-28048', 'TX1021-29092', 
                        'TX1021-35145', 'TX1021-32679', 'TX1021-23485', 'TX1021-29721', 'TX1021-28748', 'TX1021-31291-1', 'TX1021-16761', 'TX1021-24440-1', 
                        'TX1021-21594', 'TX1021-28774', 'TX1021-36917', 'TX1021-36209', 'TX1021-23349', 'TX1021-32908', 'TX1021-32237', 'TX1021-29389', 
                        'TX1021-36891', 'TX1021-16242', 'TX1021-19339', 'TX1021-29008', 'TX1021-34011-1', 'TX1021-31947', 'TX1021-28381', 'TX1021-27614', 
                        'TX1021-30620', 'TX1021-27893', 'TX1021-19249', 'TX1021-34667', 'TX1021-29631', 'TX1021-20781', 'TX1021-27769', 'TX1021-18441', 
                        'TX1021-35022', 'TX1021-17591', 'TX1021-29670', 'TX1021-27815-1', 'TX1021-28944', 'TX1021-25366', 'TX1021-28633', 'TX1021-25640', 
                        'TX1021-27093', 'TX1021-29791', 'TX1021-16468', 'TX1021-24641', 'TX1021-23148', 'TX1021-24441', 'TX1021-28875-1', 'TX1021-29983', 
                        'TX1021-22788', 'TX1021-32909-1', 'TX1021-20300', 'TX1021-19589', 'TX1021-35997', 'TX1021-29517', 'TX1021-30303', 'TX1021-30955', 
                        'TX1021-21714', 'TX1021-18965', 'TX1021-19134', 'TX1021-30428', 'TX1021-36738', 'TX1021-24941', 'TX1021-23216', 'TX1021-18176',
                        'TX1021-31863', 'TX1021-28874', 'TX1021-28031', 'TX1021-22711', 'TX1021-16739', 'TX1021-22847', 'TX1021-17895', 'TX1021-21489', 
                        'TX1021-21477', 'TX1021-32975', 'TX1021-33648', 'TX1021-33017', 'TX1021-28578', 'TX1021-32056', 'TX1021-23275', 'TX1021-34037', 'TX1021-37545', 'TX1021-29849', 
                        'TX1021-31410', 'TX1021-32767', 'TX1021-29762', 'TX1021-27393', 'TX1021-30619', 'TX1021-32057', 'TX1021-30302', 'TX1021-29630', 'TX1021-29764-1', 'TX1021-23143', 'TX1021-20760', 'TX1021-28969',
                        'TX1021-28177', 'TX1021-23310', 'TX1021-27568', 'TX1021-32678', 'TX1021-16937', 'TX1021-20355', 'TX1021-25603', 'TX1021-33644', 
                        'TX1021-16532', 'TX1021-32928', 'TX1021-31293', 'TX1021-20847', 'TX1021-24954', 'TX1021-27558', 'TX1021-28104', 'TX1021-34961', 
                        'TX1021-29010-1', 'TX1021-31498', 'TX1021-32575', 'TX1021-33507', 'TX1021-29090-1', 'TX1021-30082', 'TX1021-27651', 'TX1021-23301',
                        'TX1021-20202', 'TX1021-32066', 'TX1021-25367', 'TX1021-18485', 'TX1021-17616', 'TX1021-29999', 'TX1021-29222', 'TX1021-31946',
                        'TX1021-32236', 'TX1021-30349', 'TX1021-18342', 'TX1021-21569', 'TX1021-23302', 'TX1021-28773', 'TX1021-28981', 'TX1021-29407', 
                        'TX1021-18277', 'TX1021-29645', 'TX1021-25639', 'TX1021-31497', 'TX1021-23718', 'TX1021-32448', 'TX1021-18484', 'TX1021-30693', 
                        'TX1021-24350', 'TX1021-16185', 'TX1021-19114', 'TX1021-23683', 'TX1021-23385', 'TX1021-28577', 'TX1021-33002', 'TX1021-17450',
                              'TX1021-20580', 'TX1021-23015', 'TX1021-29925', 'TX1021-18177', 'TX1021-22013', 'TX1021-33057', 'TX1021-29850', 'TX1021-18214', 
                              'TX1021-33508', 'TX1021-31289', 'TX1021-16400', 'TX1021-31747', 'TX1021-31412-1', 'TX1021-25190', 'TX1021-33817-1', 'TX1021-24351', 
                              'TX1021-20299', 'TX1021-34793-1', 'TX1021-31019', 'TX1021-35847-1', 'TX1021-24929', 'TX1021-24147', 'TX1021-37346', 'TX1021-19113', 'TX1021-27610', 'TX1021-33064-1', 
                              'TX1021-37544', 'TX1021-29734', 'TX1021-31411', 'TX1021-27189', 'TX1021-30525', 'TX1021-24077-1', 'TX1021-17965', 'TX1021-23014', 'TX1021-29223', 'TX1021-31221', 'TX1021-29278', 'TX1021-30203', 'TX1021-24955', 'TX1021-18680', 'TX1021-29018', 'TX1021-31494', 'TX1021-21307', 'TX1021-29505', 'TX1021-23274', 'TX1021-37214', 'TX1021-28133', 'TX1021-18317', 'TX1021-24884', 'TX1021-17896', 'TX1021-29459', 'TX1021-28801', 'TX1021-23973', 'TX1021-24930', 'TX1021-18213', 'TX1021-24339', 'TX1021-17682', 'TX1021-21308', 'TX1021-32546', 'TX1021-22012', 'TX1021-28105', 'TX1021-24959-1', 'TX1021-34004', 'TX1021-23001', 'TX1021-29177', 'TX1021-24003', 'TX1021-29390', 'TX1021-29988', 'TX1021-24808', 'TX1021-24885', 'TX1021-22676', 'TX1021-34211', 'TX1021-27813', 'TX1021-29669', 'TX1021-30650', 'TX1021-33003', 'TX1021-18603', 'TX1021-27557', 'TX1021-24442', 'TX1021-16323', 'TX1021-17341', 'TX1021-21488', 'TX1021-25775', 'TX1021-28982', 'TX1021-35146-1', 'TX1021-22193', 'TX1021-23046', 'TX1021-29935', 'TX1021-33922']
voided_agreement_header = ["Void Agreements"]

void_agreements_sheet.append(voided_agreement_header)
for a in voided_agreements:
    void_agreements_sheet.append([a])

cancel_agreements =  ['TX1021-22594', 'TX1021-23368-1', 'TX1021-21510', 'TX1021-30318-1', 'TX1021-27615-1', 'TX1021-21479', 'TX1021-22713-1', 'TX1021-26497-1', 'TX1021-21597', 
                      'TX1021-30507', 'TX1021-20129', 'TX1021-13825', 'TX1021-20210', 'TX1021-21523', 'TX1021-22595', 'TX1021-23779-1', 'TX1021-24005-1', 'TX1021-21309', 'TX1021-27278-1', 'TX1021-20301', 'TX1021-14793', 'TX1021-23276-1', 
                      'TX1021-23387-1', 'TX1021-22790-1', 'TX1021-23218-1', 'TX1021-27095-1', 'TX1021-20207', 'TX1021-24149-1', 'TX1021-23975-1', 'TX1021-20727', 'TX1021-20200', 'TX1021-22590', 
                      'TX1021-23739-1', 'TX1021-30317-1', 'TX1021-22361', 'TX1021-21715', 'TX1021-21525', 'TX1021-23048-1', 'TX1021-20156', 'TX1021-18318', 'TX1021-20226', 'TX1021-22591', 'TX1021-25059-1',
                      'TX1021-21637', 'TX1021-23248-1', 'TX1021-25192-1', 'TX1021-20690', 'TX1021-27771-1', 'TX1021-20193', 'TX1021-23149-1', 'TX1021-20849', 'TX1021-20213', 'TX1021-21302', 'TX1021-20357', 'TX1021-21541', 
                      'TX1021-23331-1', 'TX1021-30321-1', 'TX1021-20214', 'TX1021-20206', 'TX1021-20209','TX1021-21304', 'TX1021-19995', 'TX1021-22360', 'TX1021-23016-1', 'TX1021-27559-1', 'TX1021-30322-1', 'TX1021-21542', 'TX1021-21094']

cancel_agreement_header =["Cancel Agreements"]

cancel_agreements_sheet.append(cancel_agreement_header)
for a in cancel_agreements:
    cancel_agreements_sheet.append([a])

workbook.save("./voided_&_cancel_agreements_list.xlsx")